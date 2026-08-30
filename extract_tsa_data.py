"""
Extracts and cleans data from Stats NZ's Tourism Satellite Account
(Year ended March 2025) into normalized CSVs ready for a relational
database. Source: Stats NZ, published 3 March 2026.
"""

import openpyxl
import csv

SRC = "/mnt/user-data/uploads/tourism-satellite-account-year-ended-march-2025-tables-1-17.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------------------------------------------------------------------
# Table 1: Tourism expenditure by component, 1999-2025
# ---------------------------------------------------------------------
ws = wb["Table 1"]
rows = []
for row in ws.iter_rows(min_row=6, max_row=32, values_only=True):
    if row[0] and isinstance(row[0], int):
        rows.append(row[:6])  # year, direct_va, indirect_va, imports, gst, total_expenditure

with open("expenditure_by_component.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["year", "direct_value_added_m", "indirect_value_added_m",
                "imports_sold_to_tourists_m", "gst_paid_m", "total_expenditure_m"])
    w.writerows(rows)

# ---------------------------------------------------------------------
# Table 2: Tourism expenditure by type of tourist, 1999-2025
# columns: year, intl_exp, _, intl_pct_chg, _, dom_exp, _, dom_pct_chg, _,
#          total_exp, _, total_pct_chg, _, total_exports, _, intl_pct_of_exports
# ---------------------------------------------------------------------
ws = wb["Table 2"]
rows = []
for row in ws.iter_rows(min_row=5, max_row=31, values_only=True):
    if row[0] and isinstance(row[0], int):
        year = row[0]
        intl_exp = row[1]
        dom_exp = row[5]
        total_exp = row[9]
        total_exports = row[13]
        intl_pct_of_exports = row[15]
        rows.append([year, intl_exp, dom_exp, total_exp, total_exports, intl_pct_of_exports])

with open("expenditure_by_tourist_type.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["year", "international_expenditure_m", "domestic_expenditure_m",
                "total_expenditure_m", "total_exports_m", "international_pct_of_exports"])
    w.writerows(rows)

# ---------------------------------------------------------------------
# Table 5: Tourism employment, 2000-2025
# ---------------------------------------------------------------------
ws = wb["Table 5"]
rows = []
for row in ws.iter_rows(min_row=6, max_row=32, values_only=True):
    if row[0] and isinstance(row[0], int):
        rows.append(row[:4])  # year, directly_employed, indirectly_employed, total

with open("employment_totals.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["year", "directly_employed", "indirectly_employed", "total_tourism_employment"])
    w.writerows(rows)

# ---------------------------------------------------------------------
# Table 6: Direct tourism employment by industry, 2023-2025 (unpivot wide->long)
# ---------------------------------------------------------------------
ws = wb["Table 6"]
industries = ["Accommodation", "Food and beverage services", "Road, rail, and water transport",
              "Air and space transport", "Other transport, transport support, and travel and tour services",
              "Rental and hiring services", "Arts and recreation services", "Retail trade",
              "Education and training", "All non-tourism-related industries", "Total industry"]

year_rows = {2023: 9, 2024: 12, 2025: 15}  # actual data rows (1-indexed)
data = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))

rows = []
for year, row_idx in year_rows.items():
    values = data[row_idx - 1][1:12]  # 11 industry columns
    for industry, val in zip(industries, values):
        if val is not None:
            rows.append([year, industry, val])

with open("employment_by_industry.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["year", "industry", "people_employed"])
    w.writerows(rows)

# ---------------------------------------------------------------------
# Table 7: Overseas visitor arrivals by region and by purpose, 2022-2025
# ---------------------------------------------------------------------
ws = wb["Table 7"]
data = list(ws.iter_rows(min_row=1, max_row=30, values_only=True))

region_rows = []
purpose_rows = []
section = None
for row in data:
    label = row[0]
    if label == "By region of last permanent residence":
        section = "region"
        continue
    elif label == "By purpose of visit":
        section = "purpose"
        continue
    if section and label and not str(label).startswith(("1.", "2.", "3.", "4.", "5.", "6.", "Note", "Source")):
        vals_2022_2025 = row[1:5]
        if all(isinstance(v, (int, float)) for v in vals_2022_2025):
            for i, year in enumerate([2022, 2023, 2024, 2025]):
                target = region_rows if section == "region" else purpose_rows
                target.append([year, label, vals_2022_2025[i]])

with open("visitor_arrivals_by_region.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["year", "region", "arrivals"])
    w.writerows(region_rows)

with open("visitor_arrivals_by_purpose.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["year", "purpose", "arrivals"])
    w.writerows(purpose_rows)

# ---------------------------------------------------------------------
# Table 9: Guest nights by origin, 2022-2025
# ---------------------------------------------------------------------
ws = wb["Table 9"]
rows = []
for row in ws.iter_rows(min_row=5, max_row=8, values_only=True):
    if row[0] and isinstance(row[0], int):
        rows.append([row[0], row[1], row[3], row[5]])  # year, intl(000), dom(000), total(000)

with open("guest_nights.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["year", "international_guest_nights_000", "domestic_guest_nights_000", "total_guest_nights_000"])
    w.writerows(rows)

# ---------------------------------------------------------------------
# Table 18 (from the second workbook): Tourism expenditure by type of
# product and type of tourist, year ended March 2024
# ---------------------------------------------------------------------
SRC2 = "/mnt/user-data/uploads/tourism-satellite-account-year-ended-march-2025-tables-18-25.xlsx"
wb2 = openpyxl.load_workbook(SRC2, data_only=True)
ws = wb2["Table 18"]
all_rows = list(ws.iter_rows(min_row=6, max_row=38, values_only=True))

# Some product names wrap onto a second row (name-only row followed by
# the data row) - merge those before extracting.
category = None
rows = []
i = 0
while i < len(all_rows):
    row = all_rows[i]
    label = row[0]
    if label in (None, "Product"):
        i += 1
        continue
    if label in ("Tourism-characteristic products", "Tourism-related products"):
        category = label
        i += 1
        continue
    if str(label).startswith(("1.", "2.", "3.", "Note", "or releases", "Source")):
        break
    # A name-only row (no data in this row) followed by a row that
    # holds the rest of the name AND the data values.
    if row[1] is None and i + 1 < len(all_rows) and all_rows[i + 1][1] is not None:
        next_row = all_rows[i + 1]
        full_name = (str(label).strip() + " " + str(next_row[0] or "").strip()).strip()
        rows.append([category, full_name, *next_row[1:6]])
        i += 2
        continue
    if row[1] is not None:
        rows.append([category, str(label).strip(), *row[1:6]])
    i += 1

with open("expenditure_by_product.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["category", "product", "domestic_business_m", "domestic_government_m",
                "domestic_household_m", "international_m", "total_m"])
    w.writerows(rows)


print("Done. Extracted 8 CSVs from the real Stats NZ Tourism Satellite Account.")
